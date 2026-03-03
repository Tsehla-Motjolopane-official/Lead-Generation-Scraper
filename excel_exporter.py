from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule


# ── Colour constants ──────────────────────────────────────────────────────────
HEADER_BG     = "1F3864"   # dark blue
HEADER_FG     = "FFFFFF"   # white
ROW_ALT       = "F2F2F2"   # light grey
ROW_WHITE     = "FFFFFF"
STAR_GREEN    = "C6EFCE"   # ≥ 4.5
STAR_YELLOW   = "FFEB9C"   # 3.5 – 4.4
STAR_RED      = "FFC7CE"   # < 3.5
STAR_GREEN_FG = "276221"
STAR_YELLOW_FG = "9C6500"
STAR_RED_FG   = "9C0006"


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=HEADER_BG)


def _alt_fill(row_idx: int) -> PatternFill:
    color = ROW_ALT if row_idx % 2 == 0 else ROW_WHITE
    return PatternFill("solid", fgColor=color)


def _apply_header(ws, headers: list[str]) -> None:
    """Write bold, coloured header row and freeze it."""
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.row_dimensions[1].height = 30


def _auto_size_columns(ws) -> None:
    """Set column widths based on the longest cell value, capped at 60."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _parse_hours(opening_hours: dict | None) -> list[str]:
    """Return list of 7 day-hour strings [Mon…Sun], empty string if unavailable."""
    if not opening_hours:
        return [""] * 7
    weekday_text = opening_hours.get("weekday_text", [])
    # weekday_text is [Mon, Tue, Wed, Thu, Fri, Sat, Sun] from the API
    days = []
    for entry in weekday_text:
        # Strip the day prefix ("Monday: 9:00 AM – 5:00 PM" → "9:00 AM – 5:00 PM")
        if ":" in entry:
            days.append(entry.split(": ", 1)[1])
        else:
            days.append(entry)
    # Pad to 7 if needed
    while len(days) < 7:
        days.append("")
    return days[:7]


def _add_star_conditional_formatting(ws, col_letter: str, last_row: int) -> None:
    """Apply traffic-light colours to the star rating column."""
    cell_range = f"{col_letter}2:{col_letter}{last_row}"

    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["4.5"],
            fill=PatternFill("solid", fgColor=STAR_GREEN),
            font=Font(color=STAR_GREEN_FG),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="between",
            formula=["3.5", "4.49"],
            fill=PatternFill("solid", fgColor=STAR_YELLOW),
            font=Font(color=STAR_YELLOW_FG),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThan",
            formula=["3.5"],
            fill=PatternFill("solid", fgColor=STAR_RED),
            font=Font(color=STAR_RED_FG),
        ),
    )


def export(businesses: list[dict], category: str, city: str, output_path: str) -> None:
    """
    Build the Excel workbook from enriched business dicts and write to output_path.

    Each business dict is expected to have the keys returned by
    google_places.get_place_details(), plus 'place_id' from the search results.
    """
    wb = Workbook()

    # ── Sheet 1: Businesses ───────────────────────────────────────────────────
    ws_biz = wb.active
    ws_biz.title = "Businesses"

    biz_headers = [
        "Business Name", "Category", "Star Rating", "Total Reviews",
        "Phone Number", "Address", "City",
        "Website", "Google Maps URL",
        "Mon Hours", "Tue Hours", "Wed Hours", "Thu Hours",
        "Fri Hours", "Sat Hours", "Sun Hours",
    ]
    _apply_header(ws_biz, biz_headers)

    for row_idx, biz in enumerate(businesses, start=2):
        hours = _parse_hours(biz.get("opening_hours"))
        row_data = [
            biz.get("name", ""),
            category,
            biz.get("rating", ""),
            biz.get("user_ratings_total", ""),
            biz.get("formatted_phone_number", ""),
            biz.get("formatted_address", ""),
            city,
            biz.get("website", ""),
            biz.get("url", ""),
            *hours,
        ]
        fill = _alt_fill(row_idx)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_biz.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            # Make URL columns clickable-looking
            if col_idx in (8, 9) and value:
                cell.font = Font(color="0563C1", underline="single", name="Calibri", size=10)
            else:
                cell.font = Font(name="Calibri", size=10)

    if len(businesses) > 0:
        _add_star_conditional_formatting(ws_biz, "C", len(businesses) + 1)

    _auto_size_columns(ws_biz)

    # ── Sheet 2: Reviews ──────────────────────────────────────────────────────
    ws_rev = wb.create_sheet(title="Reviews")

    rev_headers = ["Business Name", "Reviewer Name", "Stars", "Date", "Review Text"]
    _apply_header(ws_rev, rev_headers)

    rev_row = 2
    for biz in businesses:
        biz_name = biz.get("name", "")
        reviews = biz.get("reviews", []) or []
        for review in reviews[:5]:
            # Convert epoch timestamp to readable date
            timestamp = review.get("time", 0)
            try:
                date_str = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d")
            except (OSError, ValueError, OverflowError):
                date_str = ""

            row_data = [
                biz_name,
                review.get("author_name", ""),
                review.get("rating", ""),
                date_str,
                review.get("text", ""),
            ]
            fill = _alt_fill(rev_row)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_rev.cell(row=rev_row, column=col_idx, value=value)
                cell.fill = fill
                cell.border = _thin_border()
                cell.font = Font(name="Calibri", size=10)
                if col_idx == 5:  # Review Text — wrap
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=False)
            rev_row += 1

    _auto_size_columns(ws_rev)
    # Keep review text column readable but not too wide
    ws_rev.column_dimensions["E"].width = 60

    wb.save(output_path)
